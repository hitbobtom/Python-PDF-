# asmr.one详细内容
# -*- codeing = utf-8 -*-

import os
import re  # 正则表达式，进行文字匹配`
import time
import openpyxl

up_dir_path = os.path.abspath(os.path.join(os.getcwd(), ""))
pathPrefix = up_dir_path


class Work:
    rj = None
    title = None
    imgurl = None
    date = None
    score = None
    score_peo = None
    score_num = None
    duration = None
    sales = None
    money = None
    company = None
    notes_list = None
    tag_list = None
    voice_actor_list = None
    url = None


# 读取路径文件
def readFile(path):
    list = []
    f4 = open(path, encoding='utf-8')
    # 利用循环全部读出
    while 1 > 0:
        line = f4.readline()
        if line == '':
            break
        list.append(line)

    f4.close()
    return list


# 保存数据到表格
def saveData(datalist, savepath):
    # 创建一个新的工作簿
    wb = openpyxl.Workbook()
    ws = wb.active

    # 表头
    ws['A1'].value = 'RJ号'
    ws['B1'].value = '网址'
    ws['C1'].value = '日期'
    ws['D1'].value = '声优'
    ws['E1'].value = '标题'
    ws['F1'].value = 'Tag'
    ws['G1'].value = '全年龄、带字幕'
    ws['H1'].value = '制作社团/公司'
    ws['I1'].value = '评分'
    ws['J1'].value = '评论人数'
    ws['K1'].value = '评论数量'
    ws['L1'].value = '时长'
    ws['M1'].value = '销量'
    ws['N1'].value = '金额'
    ws['O1'].value = '封面图片地址'
    ws['P1'].value = 'A'

    # 数据
    for i in range(0, len(datalist)):
        work = datalist[i]
        ws['A' + str(i + 2)].value = work.rj
        ws['B' + str(i + 2)].value = '=HYPERLINK(P' + str(i + 2) + ',I' + str(i + 2) + ')'
        ws['C' + str(i + 2)].value = work.date
        ws['D' + str(i + 2)].value = '，'.join(work.voice_actor_list)
        ws['E' + str(i + 2)].value = work.title
        ws['F' + str(i + 2)].value = '，'.join(work.tag_list)
        ws['G' + str(i + 2)].value = '，'.join(work.notes_list)
        ws['H' + str(i + 2)].value = work.company
        ws['I' + str(i + 2)].value = work.score
        ws['J' + str(i + 2)].value = work.score_peo
        ws['K' + str(i + 2)].value = work.score_num
        ws['L' + str(i + 2)].value = work.duration
        ws['M' + str(i + 2)].value = work.sales
        ws['N' + str(i + 2)].value = work.money
        ws['O' + str(i + 2)].value = work.imgurl
        ws['P' + str(i + 2)].value = work.url

    # 保存工作簿到当前目录
    wb.save(savepath)


def get_info(txt_list, index_list):
    datalist = []

    index = 0
    for i in range(0, len(index_list)):
        index = index_list[i]
        work = Work()
        # 1.RJ号、标题、封面图片地址
        line1 = txt_list[index[0]]
        RJ_CONTENT = r'<div id="(.*)" class="col-xs-12 col-sm-4 col-md-3 col-lg-2 col-xl-2"'
        rj_match = re.search(RJ_CONTENT, line1)
        if rj_match is not None:
            rj = rj_match.group(1)
            work.rj = rj
        Title_CONTENT = r'<div role="img" aria-label="Cover of (.*)" class="q-img'
        title_match = re.search(Title_CONTENT, line1)
        if title_match is not None:
            title = title_match.group(1)
            work.title = title
        ImgUrl_CONTENT = r'<img src="(.*)\?type=main" aria-hidden="true"'
        imgurl_match = re.search(ImgUrl_CONTENT, line1)
        if imgurl_match is not None:
            imgurl = imgurl_match.group(1)
            work.imgurl = imgurl
        # 2.日期
        line2 = txt_list[index[1]]
        DATE_CONTENT = r"(\d{4}-\d{2}-\d{2})"
        date_match = re.search(DATE_CONTENT, line2)
        if date_match is not None:
            date = date_match.group(1)
            work.date = date
        # 3.评分、评论人数、评论数量、时长、销量、金额、制作社团/公司、全年龄、带字幕
        line3 = txt_list[index[2]]
        SCORE_CONTENT = r'<span class="text-weight-medium text-body1 text-red">(.{1,4})</span>'
        score_match = re.search(SCORE_CONTENT, line3)
        if score_match is not None:
            score = score_match.group(1)
            work.score = score
        SCORE_PEO_CONTENT = r'<span class="text-grey">\((\d+)\)</span>'
        score_peo_match = re.search(SCORE_PEO_CONTENT, line3)
        if score_peo_match is not None:
            score_peo = score_peo_match.group(1)
            work.score_peo = score_peo
        SCORE_NUM_CONTENT = r'chat</i><span class="text-grey">\((\d+)\)</span>'
        score_num_match = re.search(SCORE_NUM_CONTENT, line3)
        if score_num_match is not None:
            score_num = score_num_match.group(1)
            work.score_num = score_num
        DURATION_CONTENT = r'schedule</i>.*\((.*)\)</span>'
        duration_match = re.search(DURATION_CONTENT, line3)
        if duration_match is not None:
            duration = duration_match.group(1)
            if 'm' in duration:
                duration = duration.replace('m', '')
                work.duration = int(duration)
            elif 'h' in duration:
                duration = duration.replace('h', '')
                work.duration = float(duration) * 60
        SALES_CONTENT = r'<span>销量: (\d+)</span>'
        sales_match = re.search(SALES_CONTENT, line3)
        if sales_match is not None:
            sales = sales_match.group(1)
            work.sales = sales
        MONEY_CONTENT = r'<span class="q-mx-sm text-weight-medium text-h6 text-red">(.*) JPY</span>'
        money_match = re.search(MONEY_CONTENT, line3)
        if money_match is not None:
            money = money_match.group(1)
            work.money = money
        COMPANY_CONTENT = r'class="text-grey ellipsis">(.*)</div><!----></div></div><div class="row items-center">'
        company_match = re.search(COMPANY_CONTENT, line3)
        if company_match is not None:
            company = company_match.group(1)
            work.company = company.strip()

        # 全年龄、带字幕
        notes_list = []
        findNOTES = re.compile(
            r'<div class="q-chip__content col row no-wrap items-center q-anchor--skip">(.{3,6})</div></div>')
        notes_list = re.findall(findNOTES, line3)
        work.notes_list = notes_list

        # 4.Tag
        tag_list = []
        start = index[3]
        end = index[4] - 1
        k = start
        if start != 0:
            if end == -1:
                end = index[5]
            while 1:
                if (k > end):
                    break
                line4 = txt_list[k]
                TAG_CONTENT = r'            (.*)$'
                tag_match = re.search(TAG_CONTENT, line4)
                if tag_match is not None:
                    tag = tag_match.group(1)
                    tag_list.append(tag)
                k += 1
        work.tag_list = tag_list
        # 5.声优
        voice_actor_list = []
        start = index[4]
        end = index[5]
        k = start
        if start != 0:
            while 1:
                if (k > end):
                    break
                line5 = txt_list[k]
                VOICE_ACTOR_CONTENT = r'            (.*)$'
                voice_actor_match = re.search(VOICE_ACTOR_CONTENT, line5)
                if voice_actor_match is not None:
                    voice_actor = voice_actor_match.group(1)
                    voice_actor_list.append(voice_actor)
                k += 1
        work.voice_actor_list = voice_actor_list
        # 6.网址
        work.url = 'https://www.asmr.one/work/' + work.rj
        datalist.append(work)

    return datalist


def get_index(txt_list):
    index_list = []
    flag = False
    start_index = 0
    end_index = 0
    main_index = 0
    date_index = 0
    tag_index = 0
    voice_actor_index = 0

    for i in range(0, len(txt_list) - 1):
        line = txt_list[i]
        next_line = txt_list[i + 1]
        if 'col-xs-12 col-sm-4 col-md-3 col-lg-2 col-xl-2' in line:
            start_index = i
        if 'q-ml-sm q-mb-xs text-subtitle1 text-weight-regular' in line:
            main_index = i
        DATE_CONTENT = r"(\d{4}-\d{2}-\d{2})"
        date_match = re.search(DATE_CONTENT, line)
        if date_match is not None:
            date_index = i
        if 'q-ma-xs' in line:
            tag_index = i + 1
        if 'q-card__actions q-mx-xs q-pa-none q-pb-xs q-card__actions--horiz row justify-start' in line:
            voice_actor_index = i + 1
        if 'col-xs-12 col-sm-4 col-md-3 col-lg-2 col-xl-2' in next_line and start_index != 0:
            end_index = i
            flag = True
        if 'q-mt-lg q-mb-xl text-h6 text-bold text-center' in next_line and start_index != 0:
            end_index = i
            flag = True
        if flag:
            index_list.append([start_index, date_index, main_index, tag_index, voice_actor_index, end_index])
            flag = False
            start_index = 0
            end_index = 0
            main_index = 0
            date_index = 0
            tag_index = 0
            voice_actor_index = 0

    return index_list


if __name__ == "__main__":  # 当程序执行时
    txt_list = readFile(pathPrefix + "\\" + "epub.txt")

    # 1.获取一个作品的起始数组下标
    index_list = get_index(txt_list)

    # 2.提取RJ信息为数组
    datalist = get_info(txt_list, index_list)

    # 3.写入Excel表格中
    savepath = pathPrefix + "\\Excel\\" + 'result-' + time.strftime("%Y%m%d%H%M%S") + '.xlsx'
    saveData(datalist, savepath)

    print("爬取完毕！")
